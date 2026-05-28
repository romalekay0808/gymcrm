from django.shortcuts import render
from django.db.models import Sum
from .models import Trainer, Client, Training, Expense, Settlement
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.conf import settings
import os
from django.db.models.functions import TruncDate, TruncMonth
import json
from django.shortcuts import render, redirect, get_object_or_404
from .forms import TrainerForm, ExpenseForm, ClientForm

@login_required
def home(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    trainings = Training.objects.all()
    expenses = Expense.objects.all()

    if not request.user.is_superuser:
        try:
            trainer = Trainer.objects.get(user=request.user)
            trainings = trainings.filter(trainer=trainer)
        except Trainer.DoesNotExist:
            trainings = Training.objects.none()

    if start_date:
        trainings = trainings.filter(date__gte=start_date)
        expenses = expenses.filter(date__gte=start_date)

    if end_date:
        trainings = trainings.filter(date__lte=end_date)
        expenses = expenses.filter(date__lte=end_date)

    total_income = trainings.aggregate(Sum("amount"))["amount__sum"] or 0
    total_expense = expenses.aggregate(Sum("amount"))["amount__sum"] or 0

    total_trainer_salary = sum([training.trainer_income() for training in trainings])
    gym_net_income = total_income - total_trainer_salary
    clean_profit = gym_net_income - total_expense

    total_clients = Client.objects.count()
    total_trainers = Trainer.objects.count()
    total_trainings = trainings.count()

    stats_by_type = {
        "membership": {
            "name": "Абонементи",
            "count": trainings.filter(training_type="membership").count(),
            "sum": trainings.filter(training_type="membership").aggregate(Sum("amount"))["amount__sum"] or 0,
        },
        "personal": {
            "name": "Персональні тренування",
            "count": trainings.filter(training_type="personal").count(),
            "sum": trainings.filter(training_type="personal").aggregate(Sum("amount"))["amount__sum"] or 0,
        },
        "split": {
            "name": "Спліт тренування",
            "count": trainings.filter(training_type="split").count(),
            "sum": trainings.filter(training_type="split").aggregate(Sum("amount"))["amount__sum"] or 0,
        },
        "single": {
            "name": "Разові заняття",
            "count": trainings.filter(training_type="single").count(),
            "sum": trainings.filter(training_type="single").aggregate(Sum("amount"))["amount__sum"] or 0,
        },
        "ubd_membership": {
            "name": "Абонементи УБД",
            "count": trainings.filter(training_type="ubd_membership").count(),
            "sum": trainings.filter(training_type="ubd_membership").aggregate(Sum("amount"))["amount__sum"] or 0,
        },
    }

    trainers_stats = []

    for trainer in Trainer.objects.all():
        trainer_trainings = trainings.filter(trainer=trainer)

        trainer_total_sum = trainer_trainings.aggregate(Sum("amount"))["amount__sum"] or 0
        trainer_salary = sum([training.trainer_income() for training in trainer_trainings])
        gym_income = trainer_total_sum - trainer_salary

        money_in_trainer_qs = trainer_trainings.filter(money_location="trainer")
        money_on_gym_card_qs = trainer_trainings.filter(money_location="gym_card")

        money_in_trainer = money_in_trainer_qs.aggregate(Sum("amount"))["amount__sum"] or 0
        money_on_gym_card = money_on_gym_card_qs.aggregate(Sum("amount"))["amount__sum"] or 0

        trainer_should_give = sum([
            training.gym_income() for training in money_in_trainer_qs
        ])

        gym_should_give = sum([
            training.trainer_income() for training in money_on_gym_card_qs
        ])

        trainer_to_gym_paid = Settlement.objects.filter(
            trainer=trainer,
            settlement_type="trainer_to_gym"
        ).aggregate(Sum("amount"))["amount__sum"] or 0

        gym_to_trainer_paid = Settlement.objects.filter(
            trainer=trainer,
            settlement_type="gym_to_trainer"
        ).aggregate(Sum("amount"))["amount__sum"] or 0

        trainer_debt = trainer_should_give - trainer_to_gym_paid
        gym_debt = gym_should_give - gym_to_trainer_paid

        final_balance = gym_debt - trainer_debt

        if final_balance > 0:
            trainer_must_pay = 0
            gym_must_pay = final_balance
        elif final_balance < 0:
            trainer_must_pay = abs(final_balance)
            gym_must_pay = 0
        else:
            trainer_must_pay = 0
            gym_must_pay = 0

        trainers_stats.append({
            "trainer": trainer,
            "total_count": trainer_trainings.count(),
            "total_sum": trainer_total_sum,
            "trainer_salary": trainer_salary,
            "gym_income": gym_income,

            "membership_count": trainer_trainings.filter(training_type="membership").count(),
            "membership_sum": trainer_trainings.filter(training_type="membership").aggregate(Sum("amount"))["amount__sum"] or 0,

            "personal_count": trainer_trainings.filter(training_type="personal").count(),
            "personal_sum": trainer_trainings.filter(training_type="personal").aggregate(Sum("amount"))["amount__sum"] or 0,

            "split_count": trainer_trainings.filter(training_type="split").count(),
            "split_sum": trainer_trainings.filter(training_type="split").aggregate(Sum("amount"))["amount__sum"] or 0,

            "single_count": trainer_trainings.filter(training_type="single").count(),
            "single_sum": trainer_trainings.filter(training_type="single").aggregate(Sum("amount"))["amount__sum"] or 0,

            "ubd_count": trainer_trainings.filter(training_type="ubd_membership").count(),
            "ubd_sum": trainer_trainings.filter(training_type="ubd_membership").aggregate(Sum("amount"))["amount__sum"] or 0,

            "money_in_trainer": money_in_trainer,
            "money_on_gym_card": money_on_gym_card,

            "trainer_should_give": trainer_should_give,
            "gym_should_give": gym_should_give,

            "trainer_to_gym_paid": trainer_to_gym_paid,
            "gym_to_trainer_paid": gym_to_trainer_paid,

            "trainer_debt": trainer_debt,
            "gym_debt": gym_debt,

            "trainer_must_pay": trainer_must_pay,
            "gym_must_pay": gym_must_pay,
            "final_balance": final_balance,
        })

    latest_trainings = trainings.order_by("-date", "-id")[:10]
    latest_expenses = expenses.order_by("-date", "-id")[:10]

    income_by_day = {}

    for training in trainings:
        day = str(training.date)

        if day not in income_by_day:
            income_by_day[day] = 0

        income_by_day[day] += float(training.amount)

    income_labels = list(income_by_day.keys())
    income_data = list(income_by_day.values())

    trainer_sales_labels = []
    trainer_sales_data = []

    for row in trainers_stats:
        trainer_sales_labels.append(row["trainer"].name)
        trainer_sales_data.append(float(row["total_sum"]))

    memberships_by_month = (
        Training.objects
        .filter(training_type__in=["membership", "ubd_membership"])
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Sum("amount"))
        .order_by("month")
    )

    membership_month_labels = [item["month"].strftime("%m.%Y") for item in memberships_by_month]
    membership_month_data = [float(item["total"]) for item in memberships_by_month]

    personal_count = trainings.filter(
        training_type="personal"
    ).count()

    membership_count = trainings.filter(
        training_type__in=["membership", "ubd_membership"]
    ).count()

    context = {
        "start_date": start_date,
        "end_date": end_date,

        "total_income": total_income,
        "total_expense": total_expense,
        "total_trainer_salary": total_trainer_salary,
        "gym_net_income": gym_net_income,
        "clean_profit": clean_profit,

        "income_labels": json.dumps(income_labels),
        "income_data": json.dumps(income_data),
        "trainer_sales_labels": json.dumps(trainer_sales_labels),
        "trainer_sales_data": json.dumps(trainer_sales_data),
        "membership_month_labels": json.dumps(membership_month_labels),
        "membership_month_data": json.dumps(membership_month_data),
        "personal_count": personal_count,
        "membership_count": membership_count,
        "total_clients": total_clients,
        "total_trainers": total_trainers,
        "total_trainings": total_trainings,

        "stats_by_type": stats_by_type,
        "trainers_stats": trainers_stats,

        "latest_trainings": latest_trainings,
        "latest_expenses": latest_expenses,
    }

    return render(request, "gym/dashboard.html", context)

@login_required
def reports(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    trainer_id = request.GET.get("trainer")

    trainings = Training.objects.all()
    expenses = Expense.objects.all()
    trainers = Trainer.objects.all()
    
    if not request.user.is_superuser:
        try:
            trainer = Trainer.objects.get(user=request.user)
            trainings = trainings.filter(trainer=trainer)
            trainers = Trainer.objects.filter(id=trainer.id)
            trainer_id = str(trainer.id)
        except Trainer.DoesNotExist:
            trainings = Training.objects.none()
            trainers = Trainer.objects.none()

    if start_date:
        trainings = trainings.filter(date__gte=start_date)
        expenses = expenses.filter(date__gte=start_date)

    if end_date:
        trainings = trainings.filter(date__lte=end_date)
        expenses = expenses.filter(date__lte=end_date)

    if trainer_id:
        trainings = trainings.filter(trainer_id=trainer_id)

    total_income = trainings.aggregate(Sum("amount"))["amount__sum"] or 0
    total_expense = expenses.aggregate(Sum("amount"))["amount__sum"] or 0
    total_trainer_salary = sum([training.trainer_income() for training in trainings])
    gym_net_income = total_income - total_trainer_salary
    clean_profit = gym_net_income - total_expense

    context = {
        "trainers": trainers,
        "selected_trainer": trainer_id,
        "start_date": start_date,
        "end_date": end_date,
        "trainings": trainings.order_by("-date"),
        "expenses": expenses.order_by("-date"),
        "total_income": total_income,
        "total_expense": total_expense,
        "total_trainer_salary": total_trainer_salary,
        "gym_net_income": gym_net_income,
        "clean_profit": clean_profit,
    }

    return render(request, "gym/reports.html", context)
@login_required
def trainers_page(request):
    if request.user.is_superuser:
        trainers = Trainer.objects.all()
    else:
        trainers = Trainer.objects.filter(user=request.user)

    return render(request, "gym/trainers.html", {
        "trainers": trainers,
    })

@login_required
def clients_page(request):
    if request.user.is_superuser:
        clients = Client.objects.all().order_by("-created_at")
    else:
        try:
            trainer = Trainer.objects.get(user=request.user)
            client_ids = Training.objects.filter(trainer=trainer).values_list("client_id", flat=True)
            clients = Client.objects.filter(id__in=client_ids).order_by("-created_at")
        except Trainer.DoesNotExist:
            clients = Client.objects.none()

    return render(request, "gym/clients.html", {
        "clients": clients,
    })

@login_required
def finances_page(request):
    trainings = Training.objects.all()
    expenses = Expense.objects.all()
    if not request.user.is_superuser:
        try:
            trainer = Trainer.objects.get(user=request.user)
            trainings = trainings.filter(trainer=trainer)
            expenses = Expense.objects.none()
        except Trainer.DoesNotExist:
            trainings = Training.objects.none()
            expenses = Expense.objects.none()

    total_income = trainings.aggregate(Sum("amount"))["amount__sum"] or 0
    total_expense = expenses.aggregate(Sum("amount"))["amount__sum"] or 0
    total_trainer_salary = sum([training.trainer_income() for training in trainings])
    gym_net_income = total_income - total_trainer_salary
    clean_profit = gym_net_income - total_expense

    return render(request, "gym/finances.html", {
        "total_income": total_income,
        "total_expense": total_expense,
        "total_trainer_salary": total_trainer_salary,
        "gym_net_income": gym_net_income,
        "clean_profit": clean_profit,
        "expenses": expenses.order_by("-date"),
    })

from .models import Membership
from datetime import date


def memberships_page(request):

    trainer_id = request.GET.get("trainer")
    status = request.GET.get("status")
    search = request.GET.get("search")

    memberships = Membership.objects.all().order_by(
        "client__name",
        "-end_date"
    )

    latest_memberships = []
    added_clients = set()

    for membership in memberships:

        if membership.client.id in added_clients:
            continue

        added_clients.add(membership.client.id)

        latest_memberships.append(membership)

    filtered = []

    for membership in latest_memberships:

        days_left = membership.days_left()

        if trainer_id:
            if str(membership.trainer.id) != trainer_id:
                continue

        if search:
            if search.lower() not in membership.client.name.lower():
                continue

        if status == "active" and days_left <= 5:
            continue

        if status == "expiring" and not (0 <= days_left <= 5):
            continue

        if status == "expired" and days_left >= 0:
            continue

        filtered.append({
            "membership": membership,
            "days_left": days_left,
        })

    trainers = Trainer.objects.all()

    context = {
        "memberships": filtered,
        "trainers": trainers,
        "selected_trainer": trainer_id,
        "selected_status": status,
        "search": search,
    }

    return render(
        request,
        "gym/memberships.html",
        context
    )

def export_reports_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Звіт"

    headers = [
        "Дата",
        "Клієнт",
        "Тренер",
        "Тип",
        "Оплата",
        "Сума",
        "ЗП тренера",
        "Залу",
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    trainings = Training.objects.all().order_by("-date")

    row_num = 2

    for training in trainings:

        sheet.cell(row=row_num, column=1).value = str(training.date)
        sheet.cell(row=row_num, column=2).value = str(training.client)
        sheet.cell(row=row_num, column=3).value = str(training.trainer)
        sheet.cell(row=row_num, column=4).value = training.get_training_type_display()
        sheet.cell(row=row_num, column=5).value = training.get_payment_type_display()
        sheet.cell(row=row_num, column=6).value = float(training.amount)
        sheet.cell(row=row_num, column=7).value = float(training.trainer_income())
        sheet.cell(row=row_num, column=8).value = float(training.gym_income())

        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="gym_report.xlsx"'

    workbook.save(response)

    return response

def export_reports_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="gym_report.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    font_path = os.path.join(
        settings.BASE_DIR,
        "gym",
        "fonts",
        "DejaVuSans.ttf"
    )

    pdfmetrics.registerFont(
        TTFont("DejaVu", font_path)
    )
    width, height = A4

    y = height - 2 * cm

    p.setFont("DejaVu", 16)
    p.drawString(2 * cm, y, "Звіт Gym CRM")
    y -= 1 * cm

    trainings = Training.objects.all().order_by("-date")
    expenses = Expense.objects.all().order_by("-date")

    total_income = sum([t.amount for t in trainings])
    total_trainer_salary = sum([t.trainer_income() for t in trainings])
    gym_income = total_income - total_trainer_salary
    total_expense = sum([e.amount for e in expenses])
    clean_profit = gym_income - total_expense

    p.setFont("DejaVu", 11)
    lines = [
        f"Загальний дохід: {total_income:.0f} грн",
        f"ЗП тренерам: {total_trainer_salary:.0f} грн",
        f"Залу після %: {gym_income:.0f} грн",
        f"Витрати: {total_expense:.0f} грн",
        f"Чистий прибуток: {clean_profit:.0f} грн",
    ]

    for line in lines:
        p.drawString(2 * cm, y, line)
        y -= 0.7 * cm

    y -= 0.5 * cm
    p.setFont("DejaVu", 13)
    p.drawString(2 * cm, y, "Trainings / Sales")
    y -= 0.7 * cm

    p.setFont("DejaVu", 9)

    for training in trainings[:40]:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("DejaVu", 9)

        line = (
            f"{training.date} | "
            f"{training.client} | "
            f"{training.trainer} | "
            f"{training.get_training_type_display()} | "
            f"{training.amount} UAH"
        )

        p.drawString(2 * cm, y, line[:120])
        y -= 0.5 * cm

    p.showPage()
    y = height - 2 * cm

    p.setFont("DejaVu", 13)
    p.drawString(2 * cm, y, "Expenses")
    y -= 0.7 * cm

    p.setFont("DejaVu", 9)

    for expense in expenses[:40]:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont("DejaVu", 9)

        line = (
            f"{expense.date} | "
            f"{expense.get_category_display()} | "
            f"{expense.amount} UAH | "
            f"{expense.comment}"
        )

        p.drawString(2 * cm, y, line[:120])
        y -= 0.5 * cm

    p.save()

    return response

def export_backup_excel(request):

    workbook = openpyxl.Workbook()

    # =========================
    # TRAINERS
    # =========================

    sheet = workbook.active
    sheet.title = "Тренери"

    headers = [
        "Ім'я",
        "Телефон",
        "Telegram ID",
    ]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    row = 2

    for trainer in Trainer.objects.all():

        sheet.cell(row=row, column=1).value = trainer.name
        sheet.cell(row=row, column=2).value = trainer.phone
        sheet.cell(row=row, column=3).value = trainer.telegram_id

        row += 1

    # =========================
    # CLIENTS
    # =========================

    sheet = workbook.create_sheet("Клієнти")

    headers = [
        "Ім'я",
        "Телефон",
        "Дата створення",
    ]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    row = 2

    for client in Client.objects.all():

        sheet.cell(row=row, column=1).value = client.name
        sheet.cell(row=row, column=2).value = client.phone
        sheet.cell(row=row, column=3).value = str(client.created_at)

        row += 1

    # =========================
    # TRAININGS
    # =========================

    sheet = workbook.create_sheet("Тренування")

    headers = [
        "Дата",
        "Клієнт",
        "Тренер",
        "Тип",
        "Оплата",
        "Сума",
    ]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    row = 2

    for training in Training.objects.all():

        sheet.cell(row=row, column=1).value = str(training.date)
        sheet.cell(row=row, column=2).value = str(training.client)
        sheet.cell(row=row, column=3).value = str(training.trainer)
        sheet.cell(row=row, column=4).value = training.get_training_type_display()
        sheet.cell(row=row, column=5).value = training.get_payment_type_display()
        sheet.cell(row=row, column=6).value = float(training.amount)

        row += 1

    # =========================
    # MEMBERSHIPS
    # =========================

    sheet = workbook.create_sheet("Абонементи")

    headers = [
        "Клієнт",
        "Тренер",
        "Назва",
        "Ціна",
        "Дата старту",
        "Дата закінчення",
    ]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    row = 2

    for membership in Membership.objects.all():

        sheet.cell(row=row, column=1).value = membership.client.name
        sheet.cell(row=row, column=2).value = membership.trainer.name
        sheet.cell(row=row, column=3).value = membership.name
        sheet.cell(row=row, column=4).value = float(membership.price)
        sheet.cell(row=row, column=5).value = str(membership.start_date)
        sheet.cell(row=row, column=6).value = str(membership.end_date)

        row += 1

    # =========================
    # EXPENSES
    # =========================

    sheet = workbook.create_sheet("Витрати")

    headers = [
        "Дата",
        "Категорія",
        "Сума",
        "Коментар",
    ]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    row = 2

    for expense in Expense.objects.all():

        sheet.cell(row=row, column=1).value = str(expense.date)
        sheet.cell(row=row, column=2).value = expense.get_category_display()
        sheet.cell(row=row, column=3).value = float(expense.amount)
        sheet.cell(row=row, column=4).value = expense.comment

        row += 1

    # =========================
    # SETTLEMENTS
    # =========================

    sheet = workbook.create_sheet("Розрахунки")

    headers = [
        "Тренер",
        "Тип",
        "Сума",
        "Дата",
        "Коментар",
    ]

    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    row = 2

    for settlement in Settlement.objects.all():

        sheet.cell(row=row, column=1).value = settlement.trainer.name
        sheet.cell(row=row, column=2).value = settlement.get_settlement_type_display()
        sheet.cell(row=row, column=3).value = float(settlement.amount)
        sheet.cell(row=row, column=4).value = str(settlement.date)
        sheet.cell(row=row, column=5).value = settlement.comment

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="gym_backup.xlsx"'
    )

    workbook.save(response)

    return response

def trainer_add(request):
    if request.method == "POST":
        form = TrainerForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("trainers")

    else:
        form = TrainerForm()

    return render(request, "gym/trainer_form.html", {
        "form": form,
        "title": "Додати тренера",
    })


def trainer_edit(request, trainer_id):
    trainer = get_object_or_404(Trainer, id=trainer_id)

    if request.method == "POST":
        form = TrainerForm(request.POST, instance=trainer)

        if form.is_valid():
            form.save()
            return redirect("trainers")

    else:
        form = TrainerForm(instance=trainer)

    return render(request, "gym/trainer_form.html", {
        "form": form,
        "title": "Редагувати тренера",
    })

def expense_add(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("finances")

    else:
        form = ExpenseForm()

    return render(request, "gym/expense_form.html", {
        "form": form,
        "title": "Додати витрату",
    })


def expense_edit(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)

    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)

        if form.is_valid():
            form.save()
            return redirect("finances")

    else:
        form = ExpenseForm(instance=expense)

    return render(request, "gym/expense_form.html", {
        "form": form,
        "title": "Редагувати витрату",
    })

def client_add(request):
    if request.method == "POST":
        form = ClientForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("clients")

    else:
        form = ClientForm()

    return render(request, "gym/client_form.html", {
        "form": form,
        "title": "Додати клієнта",
    })


def client_edit(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if request.method == "POST":
        form = ClientForm(request.POST, instance=client)

        if form.is_valid():
            form.save()
            return redirect("clients")

    else:
        form = ClientForm(instance=client)

    return render(request, "gym/client_form.html", {
        "form": form,
        "title": "Редагувати клієнта",
    })